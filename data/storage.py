"""
data/storage.py
Excel I/O functies zonder corruptie.
"""
import pandas as pd
import openpyxl
from pathlib import Path
import streamlit as st
import os

# Pad naar de single source of truth
DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_FILE = DATA_DIR / "Begroting_Willianne_Patrick_PRO_v2.xlsx"

def file_exists():
    return EXCEL_FILE.exists()

def load_all_data():
    """Laad alle relevante data uit de Excel en retourneer als dictionary voor session_state."""
    data = {}
    
    try:
        # Laad dashboard data
        dash_df = pd.read_excel(EXCEL_FILE, sheet_name="Dashboard PRO", header=None, engine="openpyxl")
        data['dashboard'] = dict(zip(dash_df.iloc[:, 0].dropna(), dash_df.iloc[:, 1].dropna()))
        
        # Laad verbouwing data
        verbouwing_df = pd.read_excel(EXCEL_FILE, sheet_name="Verbouwing begroting", header=1, engine="openpyxl")
        cols = ['Categorie', 'Post', 'Aantal', 'Eenheid', 'Kosten per eenheid (€)', 'Totaal (€)', 'Opmerking']
        verbouwing_df = verbouwing_df[[c for c in cols if c in verbouwing_df.columns]]
        verbouwing_df = verbouwing_df.dropna(subset=['Post'])
        if 'Totaal (€)' in verbouwing_df.columns:
            verbouwing_df['Totaal (€)'] = pd.to_numeric(verbouwing_df['Totaal (€)'], errors='coerce').fillna(0)
        
        # Voeg Kamer kolom toe als die niet bestaat
        if 'Kamer' not in verbouwing_df.columns:
            verbouwing_df['Kamer'] = 'Algemeen'  # Default
        
        data['verbouwing'] = verbouwing_df
        
        # Initialiseer budget verdeling als leeg dataframe
        data['budget_verdeling'] = pd.DataFrame(columns=['Kamer', 'Toegewezen Budget (€)', 'Gerealiseerd (€)', 'Beschikbaar (€)'])
        
        return data
    except Exception as e:
        st.error(f"Fout bij inlezen Excel: {e}")
        return {}

def save_verbouwing_data(df):
    """
    Schrijft de volledige verbouwing dataframe terug naar Excel.
    Overschrijft bestaande data en voegt nieuwe rijen toe indien nodig.
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Verbouwing begroting"]
        
        # Wis bestaande data vanaf rij 3 (header is rij 1-2)
        for row in range(3, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Schrijf headers (voeg Kamer toe als nodig)
        headers = ['Categorie', 'Post', 'Aantal', 'Eenheid', 'Kosten per eenheid (€)', 'Totaal (€)', 'Opmerking', 'Kamer']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_idx).value = header
        
        # Schrijf data
        for row_idx, (_, row) in enumerate(df.iterrows(), 3):
            ws.cell(row=row_idx, column=1).value = row.get('Categorie')
            ws.cell(row=row_idx, column=2).value = row.get('Post')
            ws.cell(row=row_idx, column=3).value = row.get('Aantal')
            ws.cell(row=row_idx, column=4).value = row.get('Eenheid')
            ws.cell(row=row_idx, column=5).value = row.get('Kosten per eenheid (€)')
            ws.cell(row=row_idx, column=6).value = row.get('Totaal (€)')
            ws.cell(row=row_idx, column=7).value = row.get('Opmerking')
            ws.cell(row=row_idx, column=8).value = row.get('Kamer')
        
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Kon gegevens niet opslaan in Excel: {e}")
        return False

def save_bon_to_excel(bon_data):
    """Slaat AI gegenereerde JSON bongegevens op in een nieuw 'Bonnen' tabblad."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        if "Bonnen" not in wb.sheetnames:
            ws = wb.create_sheet("Bonnen")
            ws.append(["Datum", "Leverancier", "Bedrag (€)", "Categorie", "Kamer", "Omschrijving"])
        else:
            ws = wb["Bonnen"]
            
        ws.append([
            bon_data.get("Datum", ""),
            bon_data.get("Leverancier", ""),
            bon_data.get("Bedrag", 0.0),
            bon_data.get("Categorie", ""),
            bon_data.get("Kamer", ""),
            bon_data.get("Omschrijving", "")
        ])
        
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Fout bij opslaan bon: {e}")
        return False