"""
utils/excel_handler.py  –  Hofakkers 44  v3.0

VERBETERINGEN t.o.v. v2:
- Automatische backup-rotatie naar data/backups/ (max 10) vóór elke save
- Fix cache-consistency bug: invalidate_cache() + flag in session_state
  zorgt dat init() automatisch herlaadt na een save
- Buffer-sheet (SH_BUFFER) voor YNAB envelope-overrides
- Logging naar stderr bij fouten (geen stille failures meer)
"""
import os
import shutil
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st

from utils.config import (
    EXCEL_FILE, BACKUP_DIR, MAX_BACKUPS,
    SH_DASHBOARD, SH_MAAND, SH_VERBOUWING, SH_INBOEDEL,
    SH_SPAREN, SH_HYPOTHEEK, SH_KOSTEN, SH_BUFFER,
)

logger = logging.getLogger("hofakkers44")
logger.setLevel(logging.INFO)
if not logger.handlers:
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)s: %(message)s",
                                     datefmt="%H:%M:%S"))
    logger.addHandler(h)


# ── Kolom definities ───────────────────────────────────────────────────────
VERB_COLS = [
    "Categorie", "Post", "Aantal", "Eenheid",
    "Kosten per eenheid (€)", "Totaal (€)", "Opmerking",
]
INBOEDEL_COLS = VERB_COLS[:]
MAAND_COLS = ["Categorie", "Post", "Bedrag (€)"]
KOSTEN_COLS = [
    "Datum", "Leverancier", "Bedrag (€)", "Categorie",
    "Omschrijving", "Fiscaal_Type", "Post_Ref", "Bon_Ref",
]
BUFFER_COLS = ["Categorie", "Type", "Begroot_Override (€)", "Notitie"]


# ── Helpers ────────────────────────────────────────────────────────────────

def _sf(v) -> float:
    try:
        return float(str(v or 0).replace(",", ".").replace("€", "").strip())
    except (TypeError, ValueError):
        return 0.0


def file_exists() -> bool:
    return EXCEL_FILE.exists()


# ── Gecachte Excel-loader ──────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def _load_excel_cached(file_path: str, file_mtime: float) -> dict[str, pd.DataFrame]:
    """Cache key = (path, mtime). Auto-invalidatie bij externe wijziging."""
    return pd.read_excel(file_path, sheet_name=None, engine="openpyxl", header=None)


def _get_sheets() -> dict[str, pd.DataFrame]:
    if not file_exists():
        return {}
    try:
        mtime = EXCEL_FILE.stat().st_mtime
        return _load_excel_cached(str(EXCEL_FILE), mtime)
    except PermissionError:
        st.error("❌ Excel is geopend in een ander programma. Sluit het en herlaad.")
        return {}
    except Exception as exc:
        logger.error(f"Excel read failed: {exc}")
        st.error(f"❌ Kon Excel niet lezen: {exc}")
        return {}


def invalidate_cache() -> None:
    """Gooi de cache leeg én forceer init() opnieuw te draaien bij volgende rerun."""
    _load_excel_cached.clear()
    if "_v3_loaded" in st.session_state:
        st.session_state["_v3_loaded"] = False


# ── Backup rotatie ─────────────────────────────────────────────────────────

def _make_backup() -> Path | None:
    """Maak een tijdgestempelde backup en houd er max MAX_BACKUPS."""
    if not file_exists():
        return None
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = BACKUP_DIR / f"Begroting_{ts}.xlsx"
        shutil.copy2(EXCEL_FILE, bak)
        # Rotatie: behoud nieuwste MAX_BACKUPS
        all_baks = sorted(BACKUP_DIR.glob("Begroting_*.xlsx"),
                          key=lambda p: p.stat().st_mtime, reverse=True)
        for old in all_baks[MAX_BACKUPS:]:
            try:
                old.unlink()
            except OSError as e:
                logger.warning(f"Backup kon niet worden verwijderd: {old}: {e}")
        logger.info(f"Backup gemaakt: {bak.name}")
        return bak
    except Exception as e:
        logger.error(f"Backup mislukt: {e}")
        return None


def list_backups() -> list[Path]:
    """Lijst van alle backups, nieuwste eerst."""
    return sorted(BACKUP_DIR.glob("Begroting_*.xlsx"),
                  key=lambda p: p.stat().st_mtime, reverse=True)


def restore_backup(backup_path: Path) -> bool:
    """Herstel een backup als nieuwe bron-Excel. Maakt eerst backup van huidige."""
    try:
        if file_exists():
            _make_backup()  # huidige veiligstellen
        shutil.copy2(backup_path, EXCEL_FILE)
        invalidate_cache()
        logger.info(f"Backup hersteld: {backup_path.name}")
        return True
    except Exception as e:
        logger.error(f"Restore mislukt: {e}")
        st.error(f"❌ Herstellen mislukt: {e}")
        return False


# ── Atomisch schrijven ─────────────────────────────────────────────────────

def _atomic_save(wb: openpyxl.Workbook) -> bool:
    """Schrijf naar .tmp → backup huidige → vervang door .tmp."""
    tmp_path = str(EXCEL_FILE) + ".tmp"
    try:
        # 1. Maak backup van het huidige bestand
        _make_backup()
        # 2. Schrijf nieuwe versie naar tmp
        wb.save(tmp_path)
        # 3. Atomisch vervangen
        shutil.move(tmp_path, str(EXCEL_FILE))
        # 4. Cache invalidate + force reload
        invalidate_cache()
        return True
    except PermissionError:
        st.error("❌ Excel is geopend in een ander programma. Sluit het en probeer opnieuw.")
        logger.error("PermissionError bij save")
        return False
    except OSError as exc:
        st.error(f"❌ Bestand kon niet worden opgeslagen: {exc}")
        logger.error(f"OSError bij save: {exc}")
        return False
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def _write_sheet(wb: openpyxl.Workbook, name: str, df: pd.DataFrame) -> None:
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws = wb.create_sheet(name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)


def _open_wb() -> openpyxl.Workbook | None:
    if not file_exists():
        st.error("❌ Excel bestand niet gevonden.")
        return None
    try:
        return openpyxl.load_workbook(str(EXCEL_FILE))
    except PermissionError:
        st.error("❌ Excel is geopend in een ander programma.")
        return None
    except Exception as exc:
        logger.error(f"Open mislukt: {exc}")
        st.error(f"❌ Kon Excel niet openen: {exc}")
        return None


# ── Parsers ────────────────────────────────────────────────────────────────

def _parse_dashboard(df: pd.DataFrame) -> dict:
    result = {}
    if df is None or df.empty:
        return result
    for _, row in df.iterrows():
        key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if not key or key.lower() in ("nan", "none", ""):
            continue
        val = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
        if val is not None:
            result[key] = val
    return result


def _parse_begroting(sheet_name: str) -> pd.DataFrame:
    if not file_exists():
        return pd.DataFrame(columns=VERB_COLS)
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, header=1, engine="openpyxl")
    except Exception:
        return pd.DataFrame(columns=VERB_COLS)
    df.columns = [str(c).strip() for c in df.columns]
    for col in VERB_COLS:
        if col not in df.columns:
            df[col] = "" if col in ("Categorie","Post","Eenheid","Opmerking") else 0.0

    current_section = "Overig"
    rows = []
    for _, row in df.iterrows():
        cat_raw  = row["Categorie"]
        post_raw = row["Post"]
        post_str = str(post_raw).strip() if pd.notna(post_raw) else ""
        cat_str  = str(cat_raw).strip()  if pd.notna(cat_raw)  else ""

        if post_str.lower().startswith(("totaal","toaal")):
            continue
        if cat_str and not post_str:
            if not cat_str.lower().startswith("totale"):
                current_section = cat_str
            continue
        if not post_str or post_str.lower() == "nan":
            continue
        if cat_str.lower().startswith("totale"):
            continue

        n   = _sf(row["Aantal"])
        kpu = _sf(row["Kosten per eenheid (€)"])
        tot = _sf(row["Totaal (€)"]) or (n * kpu)
        rows.append({
            "Categorie":              cat_str if cat_str and cat_str.lower() != "nan" else current_section,
            "Post":                   post_str,
            "Aantal":                 n,
            "Eenheid":                str(row["Eenheid"]).strip() if pd.notna(row["Eenheid"]) else "",
            "Kosten per eenheid (€)": kpu,
            "Totaal (€)":             tot,
            "Opmerking":              str(row["Opmerking"]).strip() if pd.notna(row["Opmerking"]) else "",
        })
    return pd.DataFrame(rows, columns=VERB_COLS).reset_index(drop=True) if rows else \
           pd.DataFrame(columns=VERB_COLS)


def _parse_maand(sheets: dict) -> list[dict]:
    df = sheets.get(SH_MAAND)
    if df is None or df.empty:
        return []
    rows, huidig_cat = [], "Algemeen"
    for _, row in df.iterrows():
        col0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        col1 = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
        bedrag_raw = row.iloc[3] if len(row) > 3 else None
        try:
            bedrag = float(bedrag_raw)
        except (TypeError, ValueError):
            bedrag = None
        if col0 and col0 not in ("nan","None") and bedrag is None:
            skip = ("INKOMSTEN & TOESLAGEN","VARIABELE LASTEN & SPAREN","OVERZICHT","nan")
            if col0 not in skip:
                huidig_cat = col0
            continue
        if bedrag is not None and bedrag > 0:
            post = col1 if col1 and col1.lower() not in ("nan","none","") else col0
            if post and post.lower() not in ("nan","none",""):
                rows.append({"Categorie": huidig_cat, "Post": post, "Bedrag (€)": bedrag})
    return rows


def _parse_sparen(sheets: dict) -> dict:
    df = sheets.get(SH_SPAREN)
    if df is None or df.empty:
        return {"startwaarde":0,"maandinleg":500,"rendement":.05,"periode":30,"tabel":pd.DataFrame()}
    params, tabel_rows = {}, []
    for _, row in df.iterrows():
        k = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        v = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else None
        if "startwaarde" in k.lower():       params["startwaarde"] = _sf(v)
        elif "maandelijkse inleg" in k.lower(): params["maandinleg"] = _sf(v)
        elif "verwacht" in k.lower():           params["rendement"]  = _sf(v)
        elif "periode" in k.lower():            params["periode"]    = int(_sf(v))
        try:
            jaar = int(row.iloc[0])
            if 1 <= jaar <= 50 and len(row) >= 5:
                tabel_rows.append({
                    "Jaar": jaar,
                    "Beginwaarde (€)": _sf(row.iloc[1]),
                    "Inleg (€)":       _sf(row.iloc[2]),
                    "Rendement (€)":   _sf(row.iloc[3]),
                    "Eindwaarde (€)":  _sf(row.iloc[4]),
                })
        except (TypeError, ValueError):
            pass
    params.setdefault("startwaarde", 0)
    params.setdefault("maandinleg",  500)
    params.setdefault("rendement",   0.05)
    params.setdefault("periode",     30)
    params["tabel"] = pd.DataFrame(tabel_rows) if tabel_rows else \
                      pd.DataFrame(columns=["Jaar","Beginwaarde (€)","Inleg (€)",
                                            "Rendement (€)","Eindwaarde (€)"])
    return params


def _parse_hypotheek(sheets: dict) -> dict:
    df = sheets.get(SH_HYPOTHEEK)
    if df is None or df.empty:
        return {"params": {}, "tabel": pd.DataFrame()}
    params, rows = {}, []
    for _, row in df.iterrows():
        if len(row) > 8 and pd.notna(row.iloc[7]) and pd.notna(row.iloc[8]):
            k = str(row.iloc[7]).strip()
            v = row.iloc[8]
            if k and k.lower() not in ("nan","none"):
                try:    params[k] = float(v)
                except: params[k] = v
        try:
            m = int(row.iloc[0])
            if 1 <= m <= 360 and len(row) >= 6:
                rows.append({"Maand": m,
                             "Totale schuld (€)": _sf(row.iloc[1]),
                             "Rente (€)":         _sf(row.iloc[2]),
                             "Aflossing (€)":     _sf(row.iloc[3]),
                             "Restschuld (€)":    _sf(row.iloc[4]),
                             "Maandbedrag (€)":   _sf(row.iloc[5])})
        except (TypeError, ValueError):
            pass
    return {"params": params, "tabel": pd.DataFrame(rows) if rows else pd.DataFrame()}


def _parse_kosten(sheets: dict) -> pd.DataFrame:
    df = sheets.get(SH_KOSTEN)
    if df is None or df.empty:
        return pd.DataFrame(columns=KOSTEN_COLS)
    df = df.copy()
    if not df.empty:
        df.columns = [str(c).strip() for c in df.iloc[0]]
        df = df.iloc[1:].reset_index(drop=True)
    for col in KOSTEN_COLS:
        if col not in df.columns:
            df[col] = "" if col != "Bedrag (€)" else 0.0
    return df[KOSTEN_COLS].copy()


def _parse_buffer(sheets: dict) -> pd.DataFrame:
    """YNAB-overrides: per categorie een handmatige Begroot-aanpassing."""
    df = sheets.get(SH_BUFFER)
    if df is None or df.empty:
        return pd.DataFrame(columns=BUFFER_COLS)
    df = df.copy()
    if not df.empty:
        df.columns = [str(c).strip() for c in df.iloc[0]]
        df = df.iloc[1:].reset_index(drop=True)
    for col in BUFFER_COLS:
        if col not in df.columns:
            df[col] = "" if col != "Begroot_Override (€)" else 0.0
    return df[BUFFER_COLS].copy()


def _parse_planning(sheets: dict) -> list[dict]:
    """
    Laad Planning-sheet naar lijst van dicts.
    Retourneert [] als het sheet niet bestaat — init_planning() vult dan defaults in.
    """
    from utils.planning import PLAN_COLS
    df = sheets.get("Planning")
    if df is None or df.empty:
        return []
    df = df.copy()
    # Header staat op rij 0 als het een eerder weggeschreven DataFrame is
    first_val = str(df.iloc[0, 0]).strip().lower() if len(df) > 0 else ""
    if first_val in ("volgorde", "nr."):
        df.columns = [str(c).strip() for c in df.iloc[0]]
        df = df.iloc[1:].reset_index(drop=True)
    else:
        df.columns = [str(c).strip() for c in df.columns]
    rows = []
    for _, row in df.iterrows():
        item = {}
        for col in PLAN_COLS:
            val = row.get(col, "") if col in df.columns else ""
            item[col] = "" if (val is None or str(val).strip().lower() == "nan") else val
        # Zorg dat Volgorde een int is
        try:
            item["Volgorde"] = int(float(item.get("Volgorde", 0)))
        except (TypeError, ValueError):
            item["Volgorde"] = 0
        rows.append(item)
    return rows


# ── Hoofd-loader ───────────────────────────────────────────────────────────

def load_all_data() -> dict:
    sheets    = _get_sheets()
    return {
        "dashboard":  _parse_dashboard(sheets.get(SH_DASHBOARD)),
        "verbouwing": _parse_begroting(SH_VERBOUWING),
        "inboedel":   _parse_begroting(SH_INBOEDEL),
        "maand":      _parse_maand(sheets),
        "kosten":     _parse_kosten(sheets),
        "buffer":     _parse_buffer(sheets),
        "sparen":     _parse_sparen(sheets),
        "hypotheek":  _parse_hypotheek(sheets),
        "planning":   _parse_planning(sheets),
    }


# ── Specifieke savers ─────────────────────────────────────────────────────

def save_verbouwing(df: pd.DataFrame) -> bool:
    wb = _open_wb()
    if wb is None:
        return False
    _write_sheet(wb, SH_VERBOUWING, df)
    return _atomic_save(wb)


def save_inboedel(df: pd.DataFrame) -> bool:
    wb = _open_wb()
    if wb is None:
        return False
    _write_sheet(wb, SH_INBOEDEL, df)
    return _atomic_save(wb)


def save_dashboard(data: dict) -> bool:
    wb = _open_wb()
    if wb is None:
        return False
    df = pd.DataFrame(list(data.items()), columns=["Sleutel", "Waarde"])
    _write_sheet(wb, SH_DASHBOARD, df)
    return _atomic_save(wb)


def save_maand(rows: list) -> bool:
    wb = _open_wb()
    if wb is None:
        return False
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=MAAND_COLS)
    _write_sheet(wb, SH_MAAND, df)
    return _atomic_save(wb)


def save_kosten(df: pd.DataFrame) -> bool:
    wb = _open_wb()
    if wb is None:
        return False
    _write_sheet(wb, SH_KOSTEN, df)
    return _atomic_save(wb)


def save_buffer(df: pd.DataFrame) -> bool:
    """Sla YNAB envelope-overrides op."""
    wb = _open_wb()
    if wb is None:
        return False
    _write_sheet(wb, SH_BUFFER, df)
    return _atomic_save(wb)


def save_all_data(state: dict) -> bool:
    """
    Schrijf de volledige session_state atomisch terug naar Excel.
    Bevat: dashboard, verbouwing, inboedel, maand, kosten, buffer, planning.
    Sparen & Hypotheek zijn read-only (komen uit de originele Excel).
    """
    wb = _open_wb()
    if wb is None:
        return False

    # Dashboard
    dash_df = pd.DataFrame(
        list(state.get("dashboard", {}).items()),
        columns=["Sleutel", "Waarde"],
    )
    _write_sheet(wb, SH_DASHBOARD, dash_df)

    # Verbouwing & Inboedel
    for key, sheet in [("verbouwing", SH_VERBOUWING), ("inboedel", SH_INBOEDEL)]:
        df = state.get(key, pd.DataFrame(columns=VERB_COLS))
        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame(columns=VERB_COLS)
        _write_sheet(wb, sheet, df)

    # Maandbegroting
    maand = state.get("maand", [])
    maand_df = pd.DataFrame(maand) if maand else pd.DataFrame(columns=MAAND_COLS)
    _write_sheet(wb, SH_MAAND, maand_df)

    # Kosten
    kosten = state.get("kosten", pd.DataFrame(columns=KOSTEN_COLS))
    if not isinstance(kosten, pd.DataFrame):
        kosten = pd.DataFrame(columns=KOSTEN_COLS)
    _write_sheet(wb, SH_KOSTEN, kosten)

    # Buffer (YNAB overrides)
    buffer = state.get("buffer", pd.DataFrame(columns=BUFFER_COLS))
    if not isinstance(buffer, pd.DataFrame):
        buffer = pd.DataFrame(columns=BUFFER_COLS)
    _write_sheet(wb, SH_BUFFER, buffer)

    # Planning
    planning = state.get("planning", [])
    if planning:
        plan_df = pd.DataFrame(planning)
        _write_sheet(wb, "Planning", plan_df)

    return _atomic_save(wb)
